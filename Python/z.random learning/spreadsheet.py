# imported the openpyxl library off of pypi.org
# it is imported by pip installing by copying the command off the website
# onto the terminal
# then it can be found in "external libraries >python3.7 >site-packages"
# once there it can be used and imported
import openpyxl

# we imported a function of it from the charts section; Barhart and Reference
from openpyxl.chart import BarChart, Reference


# create a method to edit that file
def process_workbook(filename):
    # load the file into the python program via the openpyxl function "load_workbook"
    workbook = openpyxl.load_workbook(filename)

    # identify what sheet you want to use
    sheet = workbook["Sheet1"]

    # add a title for the new column
    sheet["d1"].value = "corrected price 2"

    # for loop starting from row 2 to row whatever with data in (remember that the last
    # number is exclusive so the "+1 is there) which puts out 2,3,4 in this case
    for row in range(2, sheet.max_row + 1):

        # choose the cell to play with (row, column) and store it in a variable
        cell = sheet.cell(row, 3)

        # store in a variable the new amount by manipulating the selected sell
        corrected_price = cell.value*0.9

        # select where the new value is going to be
        corrected_price_row = sheet.cell(row, 4)

        # paste the new value into that call
        corrected_price_row.value = corrected_price

    # save the values from column 4 into a variable
    bar_chart_values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    # save the BarChart method from "from openpyxl.chart import BarChart" at the top into a variable
    chart = BarChart()

    # add all the data to the chart
    chart.add_data(bar_chart_values)

    # choose where the top left corner will start
    sheet.add_chart(chart, "a6")

    # save that bad boy
    workbook.save(filename)


# call method
process_workbook("transactions2.xlsx")
